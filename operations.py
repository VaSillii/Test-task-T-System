import math
import os
import re
import logging
import base
import openpyxl
import time
from datetime import datetime
from openpyxl.utils.exceptions import InvalidFileException
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, WebDriverException
from selenium.webdriver.common.keys import Keys
import pyautogui


def get_data_from_browser(cities):
    data = {}
    if len(cities) == 0:
        logging.warning("In get_data_from_browser() list cities empty")
        return data
    try:
        driver = webdriver.Chrome(base.PATH_CHROME_DRIVER)
        for city in cities:
            try:
                driver.get(base.URL_SEARCH)
                elem = driver.find_element_by_name("q")
                elem.send_keys("погода " + city)
                elem.send_keys(Keys.RETURN)

                temp_c = driver.find_element_by_id('wob_tm')
                if re.match(r'^-?[0-9.,]+$', temp_c.text):
                    temp_c = float(temp_c.text)
                    data[city] = {'c': temp_c, 'f': math.ceil((temp_c * 9 / 5) + 32)}
                else:
                    temp_c = 'Not found'
                    logging.warning('Temperature for ' + city + ' not found!')
                    data[city] = {'c': temp_c, 'f': temp_c}
                assert "No results found." not in driver.page_source
            except NoSuchElementException:
                logging.warning("Error search element")
                continue
            except WebDriverException as exp:
                logging.error(exp.msg)
                break
            except Exception as exp:
                logging.error(exp)
                break
        driver.close()
    except WebDriverException as exp:
        logging.error(exp.msg)

    return data


def get_data_excel(path):
    data = []
    try:
        wb = openpyxl.load_workbook(filename=path)
        sheet = wb.active
        index_row = 3

        while sheet.cell(row=index_row, column=2).value is not None:
            if sheet.cell(row=index_row, column=2).value is None:
                break
            data.append(sheet.cell(row=index_row, column=2).value)
            index_row = index_row + 1
        wb.close()
    except InvalidFileException as exp:
        logging.error(exp)
    except KeyError as exp:
        logging.error(exp)
    return data


def save_data_excel(path_edit, path_save, data):
    try:
        wb = openpyxl.load_workbook(filename=path_edit)
        sheet = wb.active
        index_row = 3

        for key in data:
            sheet.cell(row=index_row, column=3).value = data[key]['c']
            sheet.cell(row=index_row, column=4).value = data[key]['f']
            index_row += 1

        current_time = datetime.strftime(datetime.now(), "%d.%m.%Y_%H:%M:%S")
        path_save_file = 'cities_' + current_time + '.xlsx'
        p = os.path.join(path_save, path_save_file)
        wb.save(p)
        wb.close()
    except InvalidFileException as exp:
        logging.error(exp)
    except KeyError as exp:
        logging.error(exp)
    except Exception as exp:
        logging.error(exp)


# Task 2
def open_program():
    pyautogui.hotkey('option', 'space') # correct the combination of Spotlight opening if different
    pyautogui.hotkey('option', 'space') # correct the combination of Spotlight opening if different
    time.sleep(1)
    pyautogui.typewrite(base.NAME_PROGRAM_OPENED)
    pyautogui.hotkey('enter')
    time.sleep(2)


def add_text_in_file():
    time.sleep(1)
    pyautogui.typewrite(base.POEM)


def add_modify_in_file():
    time.sleep(1)
    pyautogui.press('enter')
    time.sleep(0.5)
    pyautogui.typewrite('Lewis Carroll')
    time.sleep(1)


def remove_mouse():
    height, width = pyautogui.size()
    print(height, width)
    pyautogui.moveTo(height / 2, width / 2)


def save_as_file_poem():
    time.sleep(1)
    pyautogui.hotkey('shift', 'command', 's')
    time.sleep(3)
    current_time = datetime.strftime(datetime.now(), "%d.%m.%Y_%H:%M:%S")
    name_file = 'poem' + current_time + '.txt'
    path_file_save = os.path.join(base.PATH_DESCTOP, name_file)
    pyautogui.typewrite(path_file_save)
    time.sleep(2)
    pyautogui.hotkey('enter')
    time.sleep(2)
    pyautogui.hotkey('enter')


def exit_program():
    pyautogui.hotkey('command', 'q')

